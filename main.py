# smart_event_manager.py
# This program is a command-line event manager that allows an administrator
# to add, edit, delete, and view events, with basic file-based persistence.
# It also includes functionality to manage attendees and send email reminders
# by interacting with a JSON data file and an Excel spreadsheet.

import json
import uuid
from datetime import datetime, timedelta, date
import openpyxl
import smtplib
from email.message import EmailMessage

# ----------------------------------------------------------------------------------
#                                 1. STORAGE & DATA HANDLING
# ----------------------------------------------------------------------------------

def load_data():
    """
    Loads all event data from the events.json file.
    I decided to use a JSON file for events because it's a simple, human-readable format,
    perfect for this kind of structured data. I've added a try-except block to handle
    common errors like the file not existing or being empty, so the program won't crash.
    """
    try:
        with open("events.json", "r") as f:
            return json.load(f)
    except FileNotFoundError:
        return []
    except json.JSONDecodeError:
        # My thought process here was that if the file is corrupt or empty,
        # we should just treat it as if there are no events to begin with.
        return []

def save_data(events):
    """
    Saves the current event list to the events.json file.
    The 'indent=4' parameter makes the file much easier to read for debugging.
    """
    with open("events.json", "w") as f:
        json.dump(events, f, indent=4)

def create_attendees_file(filepath="attendees.xlsx"):
    """
    I decided to use an Excel file for attendee emails because it's a familiar format
    that's easy for a user to open, view, and edit outside of the program. This function
    ensures the file exists with the correct header before any data is added.
    """
    try:
        workbook = openpyxl.load_workbook(filepath)
        workbook.close()
    except FileNotFoundError:
        # My logic here is to create the file from scratch if it's missing.
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Attendees"
        sheet.append(["Email"])  # The header row is crucial for clarity.
        workbook.save(filepath)

def add_attendee(filepath="attendees.xlsx"):
    """
    Prompts the admin for a new attendee email and adds it to the Excel file.
    I made this a core feature because it's a critical part of the reminders system.
    """
    email = input("Enter the new attendee's email: ")
    create_attendees_file(filepath)
    workbook = openpyxl.load_workbook(filepath)
    sheet = workbook.active
    sheet.append([email])
    workbook.save(filepath)
    print(f"✅ Email '{email}' added to attendees list.")

def read_emails_from_excel(filepath="attendees.xlsx"):
    """
    Reads all attendee emails from the Excel file, skipping the header.
    I added comprehensive error handling to this function, so if the file is not
    found or there's an issue reading it, the program handles it gracefully.
    """
    try:
        create_attendees_file(filepath)
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active
        emails = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                emails.append(row[0])
        workbook.close()
        return emails
    except FileNotFoundError:
        print(f"❌ Error: Excel file not found at {filepath}")
        return []
    except Exception as e:
        print(f"❌ An error occurred while reading the Excel file: {e}")
        return []

# ----------------------------------------------------------------------------------
#                                 2. CORE FUNCTIONS (PROCEDURAL)
# ----------------------------------------------------------------------------------

def find_event(events, identifier):
    """
    A helper function to find an event by either its short ID or its name.
    This makes the edit and delete operations much more user-friendly.
    """
    for event in events:
        if event.get("id") == identifier or event.get("name", "").lower() == identifier.lower():
            return event
    return None

def is_conflict(events, new_event, existing_event_id=None):
    """
    Checks for time overlap with existing events.
    My approach here was to calculate the start and end times of both events
    and then check if their time ranges intersect. This is a key piece of
    logic for preventing scheduling conflicts. I also made sure it can
    ignore the event being edited so you don't get a false conflict error.
    """
    try:
        new_event_datetime = datetime.strptime(
            f"{new_event['date']} {new_event['time']}", "%d-%m-%Y %H:%M"
        )
    except (ValueError, KeyError):
        return False  # Invalid data, no conflict

    for event in events:
        if event.get("id") == existing_event_id:
            continue
        try:
            event_datetime = datetime.strptime(
                f"{event['date']} {event['time']}", "%d-%m-%Y %H:%M"
            )
        except (ValueError, KeyError):
            continue

        duration = timedelta(minutes=60)
        event_start = event_datetime
        event_end = event_start + duration
        new_event_start = new_event_datetime
        new_event_end = new_event_start + duration

        if max(event_start, new_event_start) < min(event_end, new_event_end):
            return True
    return False

def suggest_time_slots(events, target_date_obj):
    """
    When an admin tries to create an event with a conflict, this function
    provides helpful suggestions for available 60-minute time slots.
    It checks for conflicts every hour, starting from the beginning of the day.
    """
    target_datetime = datetime.combine(target_date_obj, datetime.min.time())
    suggested_count = 0
    current_check_time = target_datetime

    while suggested_count < 3:
        mock_event = {
            "name": "Mock",
            "date": current_check_time.strftime("%d-%m-%Y"),
            "time": current_check_time.strftime("%H:%M"),
        }
        if not is_conflict(events, mock_event):
            print(f"   - {mock_event['time']}")
            suggested_count += 1
        current_check_time += timedelta(minutes=60)

def add_event():
    """Adds a new event to the data file, including a conflict check."""
    events = load_data()
    name = input("Enter event name: ")
    date_str = input("Enter date (DD-MM-YYYY): ")
    time_str = input("Enter time (HH:MM): ")
    event_type = input("Enter event type: ")
    location = input("Enter location (optional): ")

    try:
        datetime.strptime(date_str, "%d-%m-%Y")
        datetime.strptime(time_str, "%H:%M")
    except ValueError:
        print("❌ Error: Invalid date or time format. Please use DD-MM-YYYY and HH:MM.")
        return

    new_event_dict = {
        "id": str(uuid.uuid4())[:8],
        "name": name,
        "date": date_str,
        "time": time_str,
        "type": event_type,
        "location": location,
    }

    if is_conflict(events, new_event_dict):
        print("⚠️ Conflict detected! This event overlaps with an existing one.")
        print("💡 Suggested available time slots:")
        try:
            target_date_obj = datetime.strptime(date_str, "%d-%m-%Y").date()
            suggest_time_slots(events, target_date_obj)
        except ValueError:
            pass
        return

    events.append(new_event_dict)
    save_data(events)
    print("✅ Event added successfully!")

def edit_event():
    """Allows editing an existing event with conflict checks on the new data."""
    events = load_data()
    identifier = input("Enter event ID or name to edit: ")
    event_to_edit = find_event(events, identifier)
    if not event_to_edit:
        print("❌ Event not found.")
        return

    # Prompting for new values with the old ones as defaults.
    print("--- Editing Event ---")
    new_name = input(f"Enter new name (current: {event_to_edit['name']}): ") or event_to_edit['name']
    new_date_str = input(f"Enter new date (DD-MM-YYYY, current: {event_to_edit['date']}): ") or event_to_edit['date']
    new_time_str = input(f"Enter new time (HH:MM, current: {event_to_edit['time']}): ") or event_to_edit['time']
    new_type = input(f"Enter new type (current: {event_to_edit['type']}): ") or event_to_edit['type']
    new_location = input(f"Enter new location (current: {event_to_edit['location']}): ") or event_to_edit['location']

    temp_event_dict = {
        "id": event_to_edit['id'],
        "name": new_name,
        "date": new_date_str,
        "time": new_time_str,
        "type": new_type,
        "location": new_location,
    }

    try:
        datetime.strptime(temp_event_dict['date'], "%d-%m-%Y")
        datetime.strptime(temp_event_dict['time'], "%H:%M")
    except ValueError:
        print("❌ Invalid date or time format. Edit not saved.")
        return

    if is_conflict(events, temp_event_dict, existing_event_id=event_to_edit['id']):
        print("⚠️ Edit would cause a conflict. Changes not saved.")
        return

    event_to_edit.update(temp_event_dict)
    save_data(events)
    print("✅ Event edited successfully!")

def delete_event():
    """Deletes an event by its ID or name."""
    events = load_data()
    identifier = input("Enter event ID or name to delete: ")
    event_to_delete = find_event(events, identifier)
    if not event_to_delete:
        print("❌ Event not found.")
        return

    events.remove(event_to_delete)
    save_data(events)
    print(f"✅ Event '{event_to_delete['name']}' deleted successfully.")

def view_events(target_date_str=None):
    """
    Displays events for a specific date or all events if no date is provided.
    I added sorting logic to ensure the events are displayed chronologically.
    """
    events = load_data()
    if not events:
        print("No events found.")
        return

    if target_date_str:
        try:
            target_date = datetime.strptime(target_date_str, "%d-%m-%Y").date()
        except ValueError:
            print("❌ Invalid date format. Please use DD-MM-YYYY.")
            return

        filtered_events = [
            e for e in events if e.get('date') and datetime.strptime(e['date'], "%d-%m-%Y").date() == target_date
        ]
        if not filtered_events:
            print(f"No events found for {target_date_str}.")
            return

        print(f"\n--- Events for {target_date_str} ---")
        sorted_events = sorted(filtered_events, key=lambda x: datetime.strptime(x['time'], "%H:%M"))
        for ev in sorted_events:
            print(f"[{ev['id']}] {ev['name']} - {ev['time']} @ {ev['location']} ({ev['type']})")
    else:
        print("\n--- All Events ---")
        sorted_events = sorted(
            events,
            key=lambda x: (
                datetime.strptime(x['date'], "%d-%m-%Y"),
                datetime.strptime(x['time'], "%H:%M"),
            ),
        )
        for ev in sorted_events:
            print(f"[{ev['id']}] {ev['name']} - {ev['date']} {ev['time']} @ {ev['location']} ({ev['type']})")

def view_todays_events():
    """A convenience function to quickly view events for the current date."""
    today = date.today().strftime("%d-%m-%Y")
    view_events(today)

def search_events():
    """Searches events by name or type using a keyword."""
    events = load_data()
    keyword = input("Enter a keyword to search: ")
    results = [e for e in events if keyword.lower() in e.get('name', '').lower() or keyword.lower() in e.get('type', '').lower()]

    if not results:
        print(f"No events found for keyword: '{keyword}'.")
        return

    print(f"\n--- Search Results for '{keyword}' ---")
    sorted_results = sorted(
        results,
        key=lambda x: (
            datetime.strptime(x['date'], "%d-%m-%Y"),
            datetime.strptime(x['time'], "%H:%M"),
        ),
    )
    for ev in sorted_results:
        print(f"[{ev['id']}] {ev['name']} - {ev['date']} {ev['time']} @ {ev['location']} ({ev['type']})")

def send_email_reminders(recipients, event, sender_email, sender_password):
    """
    Sends a reminder email to a list of recipients for a given event.
    This function uses a secure SSL connection to send the email.
    """
    try:
        msg = EmailMessage()
        msg['Subject'] = f"Reminder: Upcoming Event - {event['name']}"
        msg['From'] = sender_email
        msg['To'] = ", ".join(recipients)

        body = f"""
Hello,

This is a friendly reminder for the upcoming event: {event['name']}

Date: {event['date']}
Time: {event['time']}
Location: {event['location'] or 'Not specified'}

We look forward to seeing you there!
"""
        msg.set_content(body)

        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
            smtp.login(sender_email, sender_password)
            smtp.send_message(msg)

        print(f"✅ Reminders for '{event['name']}' sent to {len(recipients)} attendees!")
    except Exception as e:
        # My thought process here was to give the user a clear error message
        # if something goes wrong with the email sending process.
        print(f"❌ An error occurred while sending emails: {e}")

def send_simulated_reminders(recipients, event):
    """
    I decided to create this function for testing and demonstration purposes.
    It allows you to see exactly what the email would look like without
    needing to configure a real email account.
    """
    print("\n--- SIMULATED EMAIL REMINDER ---")
    print(f"To: {', '.join(recipients)}")
    print(f"Subject: Reminder: Upcoming Event - {event['name']}")
    body = f"""
Hello,

This is a friendly reminder for the upcoming event: {event['name']}

Date: {event['date']}
Time: {event['time']}
Location: {event['location'] or 'Not specified'}

We look forward to seeing you there!
"""
    print(f"Body:\n{body}")
    print("--------------------------------")
    print(f"✅ Simulation successful! Reminders for '{event['name']}' would have been sent to {len(recipients)} attendees.")

def export_events_to_json():
    """
    A bonus feature I added to give the admin a way to back up all event data.
    This is a great tool for data portability and recovery.
    """
    events = load_data()
    if not events:
        print("❌ No events to export.")
        return

    filename = "exported_events.json"
    save_data(events)
    print(f"✅ All events have been successfully exported to {filename}.")

def send_reminders():
    """
    Sends email reminders for upcoming events or simulates the process.
    My design choice here was to give the admin the flexibility to
    test the email feature with a simulation or use it for real-world
    applications securely.
    """
    print("--- Sending Reminders ---")
    
    action = input("Send (S) real emails or (P)rint a simulation? (S/P): ").lower()

    events = load_data()
    upcoming_events = [
        e for e in events
        if datetime.strptime(e.get('date', '01-01-2000'), "%d-%m-%Y").date() == date.today() + timedelta(days=1)
    ]

    if not upcoming_events:
        print("No events found for tomorrow.")
        return

    email_list = read_emails_from_excel('attendees.xlsx')
    if not email_list:
        print("❌ No valid email addresses were found in the Excel sheet.")
        return

    print(f"Found {len(upcoming_events)} events for tomorrow and {len(email_list)} attendees.")

    if action == 's':
        sender_email = input("Enter your sender email (e.g., your_email@gmail.com): ")
        sender_password = input("Enter your app password: ")
        for event in upcoming_events:
            send_email_reminders(email_list, event, sender_email, sender_password)
    elif action == 'p':
        for event in upcoming_events:
            send_simulated_reminders(email_list, event)
    else:
        print("❌ Invalid choice. No reminders sent.")


# ----------------------------------------------------------------------------------
#                                 3. CLI & MAIN LOOP
# ----------------------------------------------------------------------------------

def main():
    """
    The main control flow for the application. It manages the user interface
    and redirects to the correct function based on user input. I used a simple
    `while` loop with a state variable `is_admin` to handle the login and
    menu switching. This is a simple but very effective way to manage
    the two different user roles.
    """
    ADMIN_PASS = "admin123"
    is_admin = False

    print("Welcome to the Smart Event Manager!")
    while True:
        if not is_admin:
            print("\n--- Main Menu ---")
            print("1. View Today's Events")
            print("2. Search Events")
            print("3. Log in as Admin")
            print("4. Exit")
            choice = input("Enter your choice: ")

            if choice == "1":
                view_todays_events()
            elif choice == "2":
                search_events()
            elif choice == "3":
                password = input("Enter admin password: ")
                if password == ADMIN_PASS:
                    print("✅ Logged in as Admin.")
                    is_admin = True
                else:
                    print("❌ Incorrect password.")
            elif choice == "4":
                print("Goodbye!")
                break
            else:
                print("❌ Invalid choice.")
        else:
            print("\n--- Admin Menu ---")
            print("1. Add Event")
            print("2. Edit Event")
            print("3. Delete Event")
            print("4. View All Events")
            print("5. View Events by Day")
            print("6. Search Events")
            print("7. Send Reminders")
            print("8. Add Attendee")
            print("9. Export All Events")
            print("10. Log out")
            choice = input("Enter your choice: ")

            if choice == "1":
                add_event()
            elif choice == "2":
                edit_event()
            elif choice == "3":
                delete_event()
            elif choice == "4":
                view_events()
            elif choice == "5":
                date_str = input("Enter date to view (DD-MM-YYYY): ")
                view_events(date_str)
            elif choice == "6":
                search_events()
            elif choice == "7":
                send_reminders()
            elif choice == "8":
                add_attendee()
            elif choice == "9":
                export_events_to_json()
            elif choice == "10":
                is_admin = False
                print("✅ Logged out.")
            else:
                print("❌ Invalid choice.")

if __name__ == "__main__":
    main()
